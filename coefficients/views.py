from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils.translation import gettext as _
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import io

from .models import Hotel, EmissionCoefficient, EmissionCategory
from .forms import CustomLoginForm, EmissionCoefficientForm, CoefficientSearchForm


def custom_login(request):
    """Custom login view with hotel selection"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = CustomLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            hotel = form.cleaned_data.get('hotel')
            
            # Check if user is approved
            if not user.is_approved and not user.is_superuser:
                messages.error(request, _('您的账号尚未通过审批，请联系管理员'))
                return render(request, 'coefficients/login.html', {'form': form})
            
            # Update user's hotel
            user.hotel = hotel
            user.save()
            
            # Login user
            login(request, user)
            
            # Handle remember me
            if not form.cleaned_data.get('remember_me'):
                request.session.set_expiry(0)
            
            messages.success(request, _('登录成功！'))
            return redirect('dashboard')
        else:
            messages.error(request, _('用户名或密码错误'))
    else:
        form = CustomLoginForm()
    
    return render(request, 'coefficients/login.html', {'form': form})


def custom_logout(request):
    """Custom logout view"""
    logout(request)
    messages.info(request, _('您已成功退出登录'))
    return redirect('login')


@login_required
def dashboard(request):
    """Dashboard view - placeholder for future data visualization"""
    return render(request, 'coefficients/dashboard.html', {
        'user': request.user,
        'hotel': request.user.hotel
    })


@login_required
def profile(request):
    """User profile view"""
    if request.method == 'POST':
        # Update user information
        user = request.user
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.email = request.POST.get('email', '')
        user.position = request.POST.get('position', '')
        user.phone = request.POST.get('phone', '')
        user.save()
        
        messages.success(request, _('个人信息更新成功！'))
        return redirect('profile')
    
    return render(request, 'coefficients/profile.html', {
        'user': request.user
    })


def can_manage_coefficients(user):
    """Check if user can manage coefficients"""
    return user.is_authenticated and (user.can_manage_coefficients or user.is_superuser)


@login_required
@user_passes_test(can_manage_coefficients, login_url='dashboard')
def coefficient_list(request):
    """List and search emission coefficients"""
    search_form = CoefficientSearchForm(request.GET)
    coefficients = EmissionCoefficient.objects.select_related(
        'category_level1', 'category_level2', 'updated_by'
    ).all()
    
    # Search functionality
    if search_form.is_valid():
        query = search_form.cleaned_data.get('query')
        if query:
            coefficients = coefficients.filter(
                Q(product_name__icontains=query) |
                Q(product_name_en__icontains=query) |
                Q(category_level1__name__icontains=query) |
                Q(category_level2__name__icontains=query) |
                Q(coefficient__icontains=query)
            )
    
    # Sorting
    sort_by = request.GET.get('sort', '-updated_at')
    order = request.GET.get('order', 'desc')
    
    # Valid sort fields
    valid_sorts = {
        'department': 'department',
        'category_level1': 'category_level1__name',
        'category_level2': 'category_level2__name',
        'product_name': 'product_name',
        'coefficient': 'coefficient',
        'updated_at': 'updated_at',
    }
    
    if sort_by.lstrip('-') in valid_sorts:
        sort_field = valid_sorts[sort_by.lstrip('-')]
        # Toggle order if clicking the same column
        if order == 'asc':
            coefficients = coefficients.order_by(sort_field)
        else:
            coefficients = coefficients.order_by(f'-{sort_field}')
    else:
        coefficients = coefficients.order_by('-updated_at')
    
    # Pagination
    paginator = Paginator(coefficients, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get all categories for datalist
    level1_categories = EmissionCategory.objects.filter(level=1).values_list('name', flat=True)
    level2_categories = EmissionCategory.objects.filter(level=2).values_list('name', flat=True)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'level1_categories': level1_categories,
        'level2_categories': level2_categories,
        'current_sort': sort_by.lstrip('-'),
        'current_order': order,
    }
    return render(request, 'coefficients/coefficient_list.html', context)


@login_required
@user_passes_test(can_manage_coefficients, login_url='dashboard')
def coefficient_create(request):
    """Create new emission coefficient"""
    if request.method == 'POST':
        form = EmissionCoefficientForm(request.POST)
        if form.is_valid():
            coefficient = form.save(commit=False)
            coefficient.updated_by = request.user
            coefficient.save()
            messages.success(request, _('系数添加成功！'))
            return redirect('coefficient_list')
    else:
        form = EmissionCoefficientForm()
    
    # Get all categories for datalist
    level1_categories = EmissionCategory.objects.filter(level=1).values_list('name', flat=True)
    level2_categories = EmissionCategory.objects.filter(level=2).values_list('name', flat=True)
    
    context = {
        'form': form,
        'level1_categories': level1_categories,
        'level2_categories': level2_categories,
    }
    return render(request, 'coefficients/coefficient_form.html', context)


@login_required
@user_passes_test(can_manage_coefficients, login_url='dashboard')
def coefficient_edit(request, pk):
    """Edit existing emission coefficient"""
    coefficient = get_object_or_404(EmissionCoefficient, pk=pk)
    
    if request.method == 'POST':
        form = EmissionCoefficientForm(request.POST, instance=coefficient)
        if form.is_valid():
            coefficient = form.save(commit=False)
            coefficient.updated_by = request.user
            coefficient.save()
            messages.success(request, _('系数更新成功！'))
            return redirect('coefficient_list')
    else:
        form = EmissionCoefficientForm(instance=coefficient)
    
    # Get all categories for datalist
    level1_categories = EmissionCategory.objects.filter(level=1).values_list('name', flat=True)
    level2_categories = EmissionCategory.objects.filter(level=2).values_list('name', flat=True)
    
    context = {
        'form': form,
        'coefficient': coefficient,
        'level1_categories': level1_categories,
        'level2_categories': level2_categories,
    }
    return render(request, 'coefficients/coefficient_form.html', context)


@login_required
@user_passes_test(can_manage_coefficients, login_url='dashboard')
def coefficient_delete(request, pk):
    """Delete emission coefficient"""
    coefficient = get_object_or_404(EmissionCoefficient, pk=pk)
    
    if request.method == 'POST':
        coefficient.delete()
        messages.success(request, _('系数删除成功！'))
        return redirect('coefficient_list')
    
    return render(request, 'coefficients/coefficient_confirm_delete.html', {'coefficient': coefficient})


@login_required
@user_passes_test(can_manage_coefficients, login_url='dashboard')
def coefficient_export(request):
    """Export coefficients to Excel - all or selected"""
    # Check if specific IDs are provided (POST request with selected items)
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            coefficients = EmissionCoefficient.objects.filter(
                pk__in=ids
            ).select_related('category_level1', 'category_level2')
        else:
            # No IDs provided, export all
            coefficients = EmissionCoefficient.objects.select_related(
                'category_level1', 'category_level2'
            ).all()
    else:
        # GET request - export all
        coefficients = EmissionCoefficient.objects.select_related(
            'category_level1', 'category_level2'
        ).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(_("碳排放系数库"))
    
    # Define headers
    headers = [
        _('部门名称'), _('一级分类'), _('二级分类'), _('产品名称'), 
        _('单位'), _('碳排放系数'), _('最后更新'), _('特殊备注'),
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for row_num, coef in enumerate(coefficients, 2):
        ws.cell(row=row_num, column=2).value = coef.department
        ws.cell(row=row_num, column=3).value = coef.category_level1.name
        ws.cell(row=row_num, column=4).value = coef.category_level2.name
        ws.cell(row=row_num, column=5).value = coef.product_name
        ws.cell(row=row_num, column=7).value = coef.unit
        ws.cell(row=row_num, column=8).value = float(coef.coefficient)
        ws.cell(row=row_num, column=10).value = coef.updated_at.strftime('%Y/%m/%d')
        ws.cell(row=row_num, column=9).value = coef.special_note
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{_('碳排放系数库')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@user_passes_test(can_manage_coefficients, login_url='dashboard')
def coefficient_template(request):
    """Download import template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导入模板"
    
    # Define headers
    headers = [
        _('部门名称'), _('一级分类'), _('二级分类'), _('产品名称'), 
        _('单位'), _('碳排放系数'), _('特殊备注')
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add example data
    example_data = [
        ['Production', 'Seafood', 'Molluscs, other', 'Chiton（石鳖）', 'KG', '7.30', ''],
        ['R&D', 'Meat', 'Bovine meat', 'Ground Beef（牛肉末）', 'KG', '42.80', ''],
    ]
    
    for row_num, data in enumerate(example_data, 2):
        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num).value = value
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{_('碳排放系数导入模板')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
@user_passes_test(can_manage_coefficients, login_url='dashboard')
def coefficient_import(request):
    """Import coefficients from Excel"""
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    department, level1_name, level2_name, product_name, product_name_en, unit, coefficient, special_note = row[:9]
                    
                    # Validation
                    if not all([department, level1_name, level2_name, product_name, unit, coefficient]):
                        errors.append(f"第{row_num}行: 必填字段缺失")
                        error_count += 1
                        continue
                    
                    # Get or create categories
                    level1_category, created = EmissionCategory.objects.get_or_create(
                        name=level1_name,
                        level=1,
                        defaults={'parent': None}
                    )
                    
                    level2_category, created = EmissionCategory.objects.get_or_create(
                        name=level2_name,
                        level=2,
                        parent=level1_category
                    )
                    
                    # Create or update coefficient
                    EmissionCoefficient.objects.update_or_create(
                        product_name=product_name,
                        defaults={
                            'department': department,
                            'category_level1': level1_category,
                            'category_level2': level2_category,
                            'unit': unit,
                            'coefficient': float(coefficient),
                            'special_note': special_note or '',
                            'updated_by': request.user,
                        }
                    )
                    success_count += 1

                except Exception as e:
                    errors.append(f"第{row_num}行: {str(e)}")
                    error_count += 1
            
            # Show results
            if success_count > 0:
                messages.success(request, _('成功导入 %(count)s 条记录') % {'count': success_count})
            if error_count > 0:
                messages.warning(request, _('失败 %(count)s 条记录') % {'count': error_count})
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
            
        except Exception as e:
            messages.error(request, _('文件处理失败: %(error)s') % {'error': str(e)})
        
        return redirect('coefficient_list')
    
    return render(request, 'coefficients/coefficient_import.html')
